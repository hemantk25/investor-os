from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import openpyxl

STOP = {"LIMITED", "LTD", "LTD.", "INDIA", "THE", "AND", "OF", "CO", "COMPANY", "&"}


@dataclass
class ExitItem:
    stock: str
    sector: str
    category: str
    cur_value: str
    priority: str
    proceeds: str
    reason: str
    status: str = "PENDING"
    source: str = "auto"
    isin: str | None = None


@dataclass
class BuyItem:
    stock: str
    sector: str
    category: str
    alloc_lo: float
    alloc_hi: float
    target_pct: str
    conviction: str
    entry: str
    horizon: str
    thesis: str
    current_value: float = 0.0
    progress_pct: float = 0.0


@dataclass
class MonthBlock:
    label: str
    exits_text: str = ""
    buys_text: str = ""
    is_current: bool = False


@dataclass
class TargetRow:
    num: int
    stock: str
    sector: str
    status: str
    cur_value: str
    target_pct: str
    cagr: str
    conviction: str
    thesis: str


@dataclass
class Advisory:
    targets: list = field(default_factory=list)
    exits: list = field(default_factory=list)
    buys: list = field(default_factory=list)
    schedule: list = field(default_factory=list)


def _s(v):
    return "" if v is None else str(v).strip()


def _is_summary_row(stock: str) -> bool:
    return bool(re.match(r"^(ESTIMATED|TOTAL)\b", stock, re.IGNORECASE))


def _rows(ws):
    return list(ws.iter_rows(values_only=True))


def _find_header(rows, must_contain):
    for i, row in enumerate(rows):
        cells = [_s(c).lower() for c in row]
        if all(any(m in c for c in cells) for m in must_contain):
            return i, cells
    return None, None


def _col(cells, *cands):
    for j, c in enumerate(cells):
        for cand in cands:
            if c.startswith(cand):
                return j
    return None


def parse_alloc(text: str) -> tuple[float, float]:
    nums = re.findall(r"[\d.]+", text.replace("–", "-"))
    lo = float(nums[0]) if nums else 0.0
    hi = float(nums[1]) if len(nums) > 1 else lo
    mult = 1e5 if "l" in text.lower() else (1e7 if "cr" in text.lower() else 1.0)
    return lo * mult, hi * mult


def parse_advisory(path: Path) -> Advisory:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    adv = Advisory()
    for name in wb.sheetnames:
        rows = _rows(wb[name])
        low = name.lower()
        if low.startswith("target"):
            hi_, cells = _find_header(rows, ["#", "stock", "sector"])
            if hi_ is None:
                continue
            c_num, c_stock = _col(cells, "#"), _col(cells, "stock")
            c_sec, c_stat = _col(cells, "sector"), _col(cells, "status")
            c_val, c_tgt = _col(cells, "cur"), _col(cells, "target")
            c_cagr, c_conv, c_th = _col(cells, "cagr"), _col(cells, "conv"), _col(cells, "investment")
            sector = ""
            for row in rows[hi_ + 1:]:
                num = row[c_num] if c_num is not None and c_num < len(row) else None
                if isinstance(num, (int, float)):
                    adv.targets.append(TargetRow(int(num), _s(row[c_stock]), _s(row[c_sec]) or sector,
                                                 _s(row[c_stat]), _s(row[c_val]), _s(row[c_tgt]),
                                                 _s(row[c_cagr]), _s(row[c_conv]), _s(row[c_th])))
                else:
                    banner = next((_s(c) for c in row if "▶" in _s(c)), "")
                    if banner:
                        sector = banner.replace("▶", "").strip().title()
        elif low.startswith("exit"):
            hi_, cells = _find_header(rows, ["stock", "priority", "reason"])
            if hi_ is None:
                continue
            c = {k: _col(cells, k) for k in ["stock", "sector", "category", "cur", "priority", "est", "reason"]}
            for row in rows[hi_ + 1:]:
                if c["stock"] >= len(row) or not _s(row[c["stock"]]) or _is_summary_row(_s(row[c["stock"]])):
                    continue
                adv.exits.append(ExitItem(_s(row[c["stock"]]), _s(row[c["sector"]]),
                                          _s(row[c["category"]]), _s(row[c["cur"]]),
                                          _s(row[c["priority"]]), _s(row[c["est"]]),
                                          _s(row[c["reason"]])))
        elif low.startswith("new buys"):
            hi_, cells = _find_header(rows, ["stock", "allocation"])
            if hi_ is None:
                continue
            c = {k: _col(cells, k) for k in ["stock", "sector", "category", "allocation",
                                             "target", "conv", "entry", "horizon", "investment"]}
            for row in rows[hi_ + 1:]:
                if c["stock"] >= len(row) or not _s(row[c["stock"]]) or _is_summary_row(_s(row[c["stock"]])):
                    continue
                lo_a, hi_a = parse_alloc(_s(row[c["allocation"]]))
                adv.buys.append(BuyItem(_s(row[c["stock"]]), _s(row[c["sector"]]),
                                        _s(row[c["category"]]), lo_a, hi_a,
                                        _s(row[c["target"]]), _s(row[c["conv"]]),
                                        _s(row[c["entry"]]), _s(row[c["horizon"]]),
                                        _s(row[c["investment"]])))
        elif low.startswith("execution"):
            block = None
            for row in rows:
                cells = [_s(c) for c in row]
                month = next((c for c in cells if c.upper().startswith("MONTH ")), "")
                if month:
                    label = month.split("\n")[-1].strip()
                    block = MonthBlock(label=label)
                    adv.schedule.append(block)
                if block:
                    for j, c in enumerate(cells):
                        if c.upper().startswith("EXITS THIS") and j + 1 < len(cells):
                            block.exits_text = cells[j + 1]
                        if c.upper().startswith("BUY/ADD") and j + 1 < len(cells):
                            block.buys_text = cells[j + 1]
    wb.close()
    return adv


def _norm(name: str) -> str:
    toks = re.sub(r"[^A-Z0-9 ]", " ", name.upper()).split()
    return " ".join(t for t in toks if t not in STOP)


def _match_isin(adv_name: str, cons) -> str | None:
    n = _norm(adv_name)
    if not n:
        return None
    for c in cons:
        h = _norm(c.name)
        if h == n or h.startswith(n) or n.startswith(h):
            return c.isin
    return None


def apply_status(adv: Advisory, portfolio, data_dir: Path, today: date) -> Advisory:
    data_dir.mkdir(parents=True, exist_ok=True)
    cons = portfolio.consolidated()
    by_isin = {c.isin: c for c in cons}
    bl_path = data_dir / "advisory-baseline.json"
    if bl_path.exists():
        baseline = json.loads(bl_path.read_text(encoding="utf-8"))
    else:
        baseline = {"created": today.isoformat(), "exits": {}}
        for e in adv.exits:
            isin = _match_isin(e.stock, cons)
            baseline["exits"][e.stock] = {
                "isin": isin,
                "baseline_qty": by_isin[isin].qty if isin and isin in by_isin else None}
        bl_path.write_text(json.dumps(baseline, indent=2), encoding="utf-8")
    ov_path = data_dir / "rebalance-status.json"
    overrides = json.loads(ov_path.read_text(encoding="utf-8")) if ov_path.exists() else {}
    for e in adv.exits:
        if e.stock in overrides:
            e.status, e.source = overrides[e.stock], "manual"
            continue
        info = baseline["exits"].get(e.stock) or {"isin": _match_isin(e.stock, cons), "baseline_qty": None}
        e.isin = info["isin"]
        if e.isin is None:
            e.status, e.source = "REVIEW", "auto"
            continue
        cur = by_isin.get(e.isin)
        bq = info.get("baseline_qty")
        if cur is None or cur.qty <= 0:
            e.status = "DONE"
        elif bq and cur.qty <= 0.75 * bq:
            e.status = "IN PROGRESS"
        else:
            e.status = "PENDING"
    for b in adv.buys:
        isin = _match_isin(b.stock, cons)
        b.current_value = by_isin[isin].value if isin and isin in by_isin else 0.0
        b.progress_pct = min(100.0, b.current_value / b.alloc_lo * 100) if b.alloc_lo else 0.0
    for m in adv.schedule:
        try:
            d = datetime.strptime(m.label, "%b %Y")
            m.is_current = (d.year, d.month) == (today.year, today.month)
        except ValueError:
            m.is_current = False
    return adv

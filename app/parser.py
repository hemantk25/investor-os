from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl

NON_MEMBER_SHEETS = {"summary", "consolidated"}


@dataclass
class Holding:
    member: str
    name: str
    icici_symbol: str
    isin: str
    qty: float
    avg_cost: float | None
    excel_cmp: float | None
    excel_day_pct: float | None


@dataclass
class ParseResult:
    holdings: list
    skipped: list
    asof: datetime
    members: list


def _norm(cell) -> str:
    return str(cell).lower().replace("\n", " ").strip() if cell is not None else ""


def _find_header(ws):
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
        cells = [_norm(c) for c in row]
        if "isin" in cells and any(c.startswith("qty") for c in cells):
            cols = {}
            for i, c in enumerate(cells):
                if "stock name" in c:
                    cols["name"] = i
                elif c == "symbol":
                    cols["symbol"] = i
                elif c == "isin":
                    cols["isin"] = i
                elif c.startswith("qty"):
                    cols["qty"] = i
                elif c.startswith("cmp"):
                    cols["cmp"] = i
                elif c.startswith("% chg"):
                    cols["day"] = i
                elif c.startswith("avg buy"):
                    cols["avg"] = i
            return r, cols
    return None, None


def _num(v) -> float | None:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("₹", "").replace("%", "").replace("+", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _day_pct(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, str) and "%" in v:
        return _num(v)                       # "+0.65%" -> 0.65
    n = _num(v)
    return None if n is None else n * 100.0  # 0.01 fraction -> 1.0


def parse_holdings(path: Path) -> ParseResult:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    holdings, skipped, members = [], [], []
    for sheet in wb.sheetnames:
        if sheet.lower() in NON_MEMBER_SHEETS:
            continue
        ws = wb[sheet]
        hrow, cols = _find_header(ws)
        if not cols or "isin" not in cols or "qty" not in cols:
            continue
        members.append(sheet)
        for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
            name = row[cols["name"]] if cols.get("name") is not None else None
            if name is None:
                continue
            isin = str(row[cols["isin"]] or "").strip()
            qty = _num(row[cols["qty"]])
            if len(isin) != 12 or not isin.startswith("IN") or not qty or qty <= 0:
                skipped.append(f"{sheet}: {name} (missing/invalid ISIN or qty)")
                continue
            avg = _num(row[cols["avg"]]) if "avg" in cols else None
            holdings.append(Holding(
                member=sheet, name=str(name).strip(),
                icici_symbol=str(row[cols["symbol"]] or "").strip() if "symbol" in cols else "",
                isin=isin, qty=qty,
                avg_cost=avg if avg and avg > 0 else None,
                excel_cmp=_num(row[cols["cmp"]]) if "cmp" in cols else None,
                excel_day_pct=_day_pct(row[cols["day"]]) if "day" in cols else None))
    wb.close()
    asof = datetime.fromtimestamp(Path(path).stat().st_mtime)
    return ParseResult(holdings=holdings, skipped=skipped, asof=asof, members=members)

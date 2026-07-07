from pathlib import Path

from openpyxl import Workbook

HERE = Path(__file__).parent
H = ["Sr.", "Stock Name", "Symbol", "ISIN", "Qty\nHeld", "CMP (₹)", "% Chg",
     "Market\nValue (₹)", "Avg Buy\nPrice (₹)", "Total\nCost (₹)",
     "Unrealised\nGain/Loss (₹)", "Gain/\nLoss %"]


def member_sheet(wb, name, rows):
    ws = wb.create_sheet(name)
    ws.append([f"{name} – Equity Portfolio (fixture)"])
    ws.append(H)
    for r in rows:
        ws.append(r)


def holdings():
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Summary")
    ws.append(["Family Equity Portfolio (fixture)"])
    ws = wb.create_sheet("Consolidated")
    ws.append(["ignored by parser"])
    member_sheet(wb, "PK", [
        [1, "ALPHA MOTORS LIMITED", "ALPMOT", "INE001A01001", 100, 250.0, 0.01, 25000, 200.0, 20000, 5000, 0.25],
        [2, "BETA PHARMA LIMITED", "BETPHA", "INE002B01012", 50, 400.0, -0.02, 20000, None, None, None, None],
        [3, "GAMMA UNLISTED CO", "GAMUNL", "IN9999X99999", 30, 10.0, 0.0, 300, 12.0, 360, -60, -0.17],
        [4, "JUNK ROW NO ISIN", "JUNK", None, 10, 5.0, 0, 50, 5.0, 50, 0, 0],
    ])
    member_sheet(wb, "CK", [
        [1, "ALPHA MOTORS LIMITED", "ALPMOT", "INE001A01001", 50, 250.0, 0.01, 12500, 220.0, 11000, 1500, 0.14],
        [2, "DELTA BANK LIMITED", "DELBAN", "INE004D01034", 200, 150.0, 0.005, 30000, 100.0, 20000, 10000, 0.50],
    ])
    wb.save(HERE / "sample-holdings.xlsx")


def advisory():
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Target 50-Stock Portfolio")
    ws.append([])
    ws.append(["", "TARGET PORTFOLIO (fixture)"])
    ws.append([])
    ws.append(["", "#", "Stock", "Sector", "Status", "Cur. Value", "Target %", "CAGR Est.", "Conv.", "Investment Thesis"])
    ws.append(["", "  ▶  BANKING"])
    ws.append(["", 1, "Delta Bank", "Banking", "KEEP", "₹0.3L", "5.0%", "~15% pa", "★★★★★", "Fixture thesis"])
    ws.append(["", 2, "Alpha Motors", "Auto", "KEEP (TRIM)", "₹0.4L", "4.0%", "~12% pa", "★★★★", "Fixture thesis 2"])
    ws = wb.create_sheet("Exit List (25 Stocks)")
    ws.append([])
    ws.append(["", "EXIT LIST (fixture)"])
    ws.append([])
    ws.append(["", "Stock", "Sector", "Category", "Cur. Value", "G/L Est.", "Priority", "Est. Proceeds", "Reason for Exit"])
    ws.append(["", "Beta Pharma", "Pharma", "ZOMBIE", "₹0.2L", "-50%", "IMMEDIATE", "~₹20K", "Fixture reason"])
    ws.append(["", "Omega Ghost", "Misc", "ZOMBIE", "₹100", "-99%", "IMMEDIATE", "~₹100", "No match in holdings"])
    ws.append(["", "ESTIMATED TOTAL PROCEEDS — 2 EXITS", "", "", "", "", "", "~₹20K", ""])
    ws = wb.create_sheet("New Buys (10 Stocks)")
    ws.append([])
    ws.append(["", "NEW BUYS (fixture)"])
    ws.append([])
    ws.append(["", "Stock", "Sector", "Category", "Allocation", "Target%", "Conv.", "Entry Strategy", "Horizon", "Investment Thesis"])
    ws.append(["", "Delta Bank", "Banking", "Banks", "₹1–2L", "2.0%", "★★★★★", "2 tranches", "3–5yr", "Fixture buy thesis"])
    ws = wb.create_sheet("Execution Schedule")
    ws.append([])
    ws.append(["", "SCHEDULE (fixture)"])
    ws.append([])
    ws.append(["", "MONTH 1\nJul 2026", "EXITS THIS MONTH", "Exit zombies"])
    ws.append(["", "", "BUY/ADD THIS MONTH", "Start Delta Bank"])
    ws.append([])
    ws.append(["", "MONTH 2\nAug 2026", "EXITS THIS MONTH", "Trim Alpha"])
    ws.append(["", "", "BUY/ADD THIS MONTH", "Finish Delta Bank"])
    wb.save(HERE / "sample-advisory.xlsx")


if __name__ == "__main__":
    holdings()
    advisory()
    print("fixtures written")
